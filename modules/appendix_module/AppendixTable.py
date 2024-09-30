import pandas as pd
import openpyxl
from modules.Table import Table
from openpyxl.worksheet.worksheet import Worksheet
from modules.DBConnection import DBConnection

# _rows_to_skip = [14,19,25,28,39,48,49,54,59,63,67,76,81,82,91,95,97,102,107,114,115,119,124,133]
_rows_to_skip = []

# список индексов ячеек в строке с месяцами с учетом -1
_month_indices = [7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 20, 21, 25, 26, 27, 28, 29, 30, 34, 35, 36, 37, 38, 39]
_JOBS = 0
_UNITS = 1
_MONTH = 10
_PLAN_FACT = 11
_YEAR_PLAN = 2

class AppendixTable(Table):
    @staticmethod
    def job(path_to_load: str, path_to_save: str) -> None:
        try:
            appendix = openpyxl.load_workbook(path_to_load)
            requests = openpyxl.load_workbook(path_to_save)
            sheets = AppendixTable.__get_list_of_sheets(appendix)
            for sheet in sheets:
                df = pd.read_excel(path_to_load, sheet_name=sheet, engine='openpyxl')
                AppendixTable.__fill_with_one_sheet(df, requests['PO_2024_MONTH_TEMP'])
                requests.save(path_to_save)
        except PermissionError:
            pass

    @staticmethod
    def __make_row(df: pd.DataFrame, row: int, i: int, year: int, project: str, area: str,license: str, stage: str) -> list:

        work = df.iloc[row, _JOBS]
        mu = df.iloc[row, _UNITS]
        df.iloc[_MONTH, _month_indices[i - (i % 2 != 0)]]
        return [
                df.iloc[row, _JOBS], # вид работ
                df.iloc[row, _UNITS], # ед. изм
                df.iloc[_MONTH, _month_indices[i - (i % 2 != 0)]], # месяц
                year, # год
                df.iloc[_PLAN_FACT, _month_indices[i]], # план/факт
                df.iloc[row, _month_indices[i]], # 
                project, # проект
                area, # участок проекта
                license,
                stage, # стадия работ
                str(project) + '-' + str(area) # проект-участок

                ]

    @staticmethod
    def __make_row_year(df: pd.DataFrame, row: int, year: int, project: str, area: str,license: str, stage: str) -> list:
        return [df.iloc[row,_JOBS], df.iloc[row,_UNITS], year, df.iloc[row,_YEAR_PLAN], project, area, license, stage, str(project) + '-' + str(area),]
                
    @staticmethod
    def __fill_with_one_sheet(df: pd.DataFrame, sheet: Worksheet) -> None:
        project, area, license, stage, year = AppendixTable.__get_service_info(df)
        # query = DBConnection()
        for row in range(13, 141):
            if row in _rows_to_skip:
                continue
            if str(df.iloc[row,2]) != 'nan' and df.iloc[row,2] != 0:
                new_row = AppendixTable.__make_row_year(df,row,year,project,area,license,stage)
                print(new_row)
                sheet.append(new_row)
            
                # query.query_execute("INSERT [ПО_2023_ПЛАН_ГОД_СКРИПТ_ДЛЯ_ТЕСТОВ] ([Вид работ],[Ед изм],[Год],[Значение],[Проект (Паспорт ГРП)],[Участок],[Стадия],[Проект-Участок],[Видработ_едизм]) VALUES (?,?,?,?,?,?,?,?,?)",new_row)
            for i in range(0, len(_month_indices), 2):           
                plan = df.iloc[row, _month_indices[i]]

                fact = df.iloc[row, _month_indices[i + 1]]
                df = df.fillna(0)
                if (plan != 0 and str(plan) != 'nan') or (df.iloc[row,2] != 0 and str(df.iloc[row,2]) != 'nan'):
                # if (plan != 0 and str(plan) != 'nan'): # or (df.iloc[row,2] != 0 and str(df.iloc[row,2]) != 'nan'):
                    new_row = AppendixTable.__make_row(df, row, i, year, project, area, license, stage)
                    print(new_row)                    
                    # query.query_execute("INSERT INTO "+query.table_name+" ([Вид работ],[Ед изм],[Месяц],[Год],[План/Факт],[Значение],[Проект (Паспорт ГРП)],[Участок],[Стадия],[Проект-Участок],[Видработ_едизм])VALUES (?,?,?,?,?,?,?,?,?,?,?)",new_row)
                    
                    sheet.append(new_row)
                if (str(fact) != 'nan' and fact != 0) or (df.iloc[row,2] != 0 and str(df.iloc[row,2]) != 'nan'):
                # if (str(fact) != 'nan' and fact != 0): # or (df.iloc[row,2] != 0 and str(df.iloc[row,2]) != 'nan'):
                    new_row = AppendixTable.__make_row(df, row, i + 1, year, project, area, license, stage)
                    
                    # query.query_execute("INSERT INTO "+query.table_name+" ([Вид работ],[Ед изм],[Месяц],[Год],[План/Факт],[Значение],[Проект (Паспорт ГРП)],[Участок],[Стадия],[Проект-Участок],[Видработ_едизм])VALUES (?,?,?,?,?,?,?,?,?,?,?)",new_row)
                    sheet.append(new_row)

    @staticmethod
    def __get_list_of_sheets(workbook: openpyxl.Workbook) -> list:
        temp_sheets = workbook.sheetnames
        sheets = []
        for i in range(temp_sheets.index('(1)') + 1, temp_sheets.index('(2)')):
            sheets.append(temp_sheets[i])
        return sheets

    @staticmethod
    def __get_service_info(df: pd.DataFrame) -> list[str, str, str, str, int]:
        column = df[df.columns[1]].values
        project = column[1]
        area = column[3]
        stage = column[5]
        license = column[4]
        print(stage)
        year = df[df.columns[2]].values[10].split(' ')[1]
        return project, area, license, stage, int(year)
    


