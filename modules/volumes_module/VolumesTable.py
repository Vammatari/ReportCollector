import openpyxl
import re
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from modules.Table import Table
from modules.Settings import Settings
_sheet_reg = r'\d{2}-\d{4}'
_year_reg = r'\d{4}'
_month_map = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель', '05': 'Май', '06': 'Июнь',
                '07': 'Июль', '08': 'Август', '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}
_ve_list = [
'Бурение колонковое с поверхности',
'Бурение из подземных горных выработок',
'Бурение колонковое МГБУ',
'Бурение КГК',
'RC-бурение',
'Бурение бескерновое',
'Бурение технологических скважин',
'Бурение гидрогеологическое',
'Бурение инженерно-геологическое',
'Проходка поверхностных горных выработок (механизированная)',
'Проходка поверхностных горных выработок (ручная)',
'Проходка подземных горных выработок',
'Геохимические работы - мелкометражное бурение',
'Геохимические работы по ПОР',
'Геохимические работы по ВОР',
'Геохимические работы по ПР',
'Наземные геофизические исследования профильные',
'Наземные геофизические исследования площадные',
'Аэрогеофизические исследования',
'Аэрогеофизические исследования БПЛА',
'Прочие виды геофизических исследований',
'Пробирный анализ',
'Атомно-абсорбционный анализ',
'Рентген-флуоресцентный анализ',
'Другие виды анализов',
'Технологические исследования',
'Тахеометрическая съемка',
'Топографо-геодезические работы',
'Геологические маршруты',
'Документирование керна скважин и поверхности горных выработок',
'Отбор проб',
'Отбор технологических проб',
'Пробоподготовка',
'Петрографические и минералогические исследования',
'Камеральные работы',
'Транспортные расходы',
'Прочие затраты'
]

class VolumesTable(Table):
    @staticmethod
    def job(path_to_load: str, path_to_save: str) -> None:
        """ - """
        try:
            # загружаем книгу
            volumes = openpyxl.load_workbook(path_to_load)
            # формируем отчет
            result = openpyxl.load_workbook(path_to_save)
            # формируем список листов книги с месяцами
            sheets = VolumesTable.__get_list_of_sheets(volumes)
            
            for sheet in sheets:

                df = pd.read_excel(path_to_load, sheet_name=sheet, engine='openpyxl')
                VolumesTable.__fill_with_one_sheet(df, result['Объёмы-Затраты'], VolumesTable.__get_month(sheet))
                result.save(path_to_save)
        except PermissionError:
            pass

    @staticmethod
    def __get_list_of_sheets(workbook: openpyxl.Workbook) -> list:
        """На вход поступает книга Excel\n
        Возвращаем список имён листов удовлетворяющих условия отбора """
        return [sheet for sheet in workbook.sheetnames if re.match(_sheet_reg, sheet) is not None]

    @staticmethod
    def __get_service_info(df: pd.DataFrame) -> list[str, str]:
        """ забираем название филиала и год(в объединённых ячейках видит только адрес первой) """
        branch = df[df.columns[3]].values[1]
        year = df.iloc[3, VolumesTable.__get_length(df) - 1]
        return branch, year

    @staticmethod
    def __get_month(sheet: str) -> str:
        """из названия листа забираем первые два символа - индекс месяца и возвращаем ключ(название месяца)"""
        return _month_map.get(sheet[0:2])

    @staticmethod
    def __get_length(df: pd.DataFrame) -> int:
        # возвращаем ширину таблицы
        col = 3
        while re.match(_year_reg, str(df.iloc[3, col])) is None:
            col += 1
        # print(df.iloc[3,col])    
        return col + 1

    @staticmethod
    def __get_column_info(df: pd.DataFrame, col_number: int) -> list[str, str, str]:
        organization = df.iloc[3, col_number]
        project = df.iloc[4, col_number]
        area = df.iloc[5, col_number]
        license = df.iloc[6,col_number]
        organization1c = df.iloc[7,col_number]
        # if license == 0:
        #     license = ''
        # if project == 0:
        #     project = ''
        # if area == 0:
        #     area = ''
        # if organization1c == 0:
        #     organization1c = ''
        return organization, project, area, license, organization1c

    @staticmethod
    def __fill_with_one_sheet(df: pd.DataFrame, sheet: Worksheet, month: str) -> None:
        branch, year = VolumesTable.__get_service_info(df)
        work_type = ''
        plan_fact = ''
        count = 0
        for row in range(8, 196):
            work_type_temp = df.iloc[row, 0]
            work_type = work_type_temp if str(work_type_temp) != 'nan' else work_type
            plan_fact_temp = df.iloc[row, 1]
            plan_fact = plan_fact_temp if str(plan_fact_temp) != 'nan' else plan_fact
            unit = df.iloc[row, 2]
            length = VolumesTable.__get_length(df)
            for col in range(3, length - 1):
                amount = df.iloc[row,col]
                if str(amount) in ('nan','0',''):
                    continue
                # if df.iloc[row,col] != 0 and str(df.iloc[row,col]) != 'nan':
                #     if str(df.iloc[row,col]).strip() in ['Примечания:','Лицензии к получению в 2023 году']:
                #         continue
                organization, project, area, license, organization1c = VolumesTable.__get_column_info(df, col)
                if str(organization) in ('nan','','Организация','0'):
                    continue
                new_row = [work_type, plan_fact, unit, amount, organization, project, area, branch, license, organization1c, month, year]
                print(new_row)
                sheet.append(new_row)
        # count = 0
        # for row in range(1,196):
        #     work_type = df.iloc[row,0]

        #     if work_type in _ve_list:
        #         count +=1
        #         print(True,work_type,count)
